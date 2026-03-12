"""Generador de reportes Excel."""
from typing import Dict, Any
from io import BytesIO
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel_report(audit_data: Dict[str, Any]) -> BytesIO:
    """
    Genera un reporte de auditoría en formato Excel (.xlsx).
    
    Args:
        audit_data: Diccionario con datos de la auditoría
            - paciente: Datos del paciente
            - riesgo_global: Nivel de riesgo (alto/medio/bajo)
            - total_hallazgos: Total de hallazgos
            - exposicion_glosas: Exposición económica en COP
            - hallazgos: Lista de hallazgos
            - hallazgos_por_riesgo: Conteo por riesgo
            - hallazgos_por_modulo: Conteo por módulo
            - recomendacion_general: Recomendación general
    
    Returns:
        BytesIO con contenido del archivo Excel
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Instalar con: pip install openpyxl")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Auditoría"
    
    # Estilos
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    title_font = Font(bold=True, size=12)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colores según riesgo
    risk_colors = {
        "alto": "FEF2F2",
        "medio": "FFFBEB",
        "bajo": "F0FDF4",
        "pending": "F3F4F6"
    }
    
    paciente = audit_data.get("paciente", {})
    riesgo = audit_data.get("riesgo_global", "bajo")
    hallazgos = audit_data.get("hallazgos", [])
    
    # ===== ENCABEZADO =====
    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = "🏥 REPORTE DE AUDITORÍA MÉDICA"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(italic=True, size=10, color="666666")
    
    row += 2
    
    # ===== INFORMACIÓN DEL PACIENTE =====
    ws[f'A{row}'] = "INFORMACIÓN DEL PACIENTE"
    ws[f'A{row}'].font = title_font
    row += 1
    
    # Tabla de información
    info_data = [
        ["Identificador", paciente.get('label', 'N/A')],
        ["Diagnóstico Principal", paciente.get('diagnostico_principal', 'N/A')],
        ["Código CIE-10", paciente.get('codigo_cie10', 'N/A')],
        ["Días Hospitalización", paciente.get('dias_hospitalizacion', 'N/A')],
        ["Nivel de Riesgo", riesgo.upper()],
        ["Total Hallazgos", audit_data.get('total_hallazgos', 0)],
        ["Exposición a Glosas", f"${audit_data.get('exposicion_glosas', 0):,.0f} COP"],
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border_thin
        ws[f'B{row}'].border = border_thin
        
        # Color especial para riesgo
        if label == "Nivel de Riesgo":
            ws[f'B{row}'].fill = PatternFill(
                start_color=risk_colors.get(riesgo, "F3F4F6"),
                end_color=risk_colors.get(riesgo, "F3F4F6"),
                fill_type="solid"
            )
            ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
    
    row += 2
    
    # ===== HALLAZGOS =====
    ws[f'A{row}'] = "HALLAZGOS DE AUDITORÍA"
    ws[f'A{row}'].font = title_font
    row += 1
    
    # Encabezados de tabla
    headers = ["Módulo", "Riesgo", "Descripción", "Recomendación", "Valor Glosa", "Normativa", "Estado"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    row += 1
    
    # Datos de hallazgos
    if hallazgos:
        for finding in hallazgos:
            modulo = finding.modulo if hasattr(finding, 'modulo') else finding.get('modulo', 'N/A')
            riesgo_f = finding.riesgo if hasattr(finding, 'riesgo') else finding.get('riesgo', 'bajo')
            desc = finding.descripcion if hasattr(finding, 'descripcion') else finding.get('descripcion', '')
            rec = finding.recomendacion if hasattr(finding, 'recomendacion') else finding.get('recomendacion', '')
            valor = finding.valor_glosa_estimado if hasattr(finding, 'valor_glosa_estimado') else finding.get('valor_glosa_estimado', 0)
            normativa = finding.normativa_aplicable if hasattr(finding, 'normativa_aplicable') else finding.get('normativa_aplicable', '')
            estado = finding.estado if hasattr(finding, 'estado') else finding.get('estado', 'activo')
            
            ws[f'A{row}'] = modulo.replace('_', ' ').title()
            ws[f'B{row}'] = riesgo_f.upper()
            ws[f'C{row}'] = desc
            ws[f'D{row}'] = rec
            ws[f'E{row}'] = f"${valor:,.0f}" if valor else "N/A"
            ws[f'F{row}'] = normativa if normativa else "N/A"
            ws[f'G{row}'] = estado.capitalize()
            
            # Aplicar color según riesgo
            row_fill = PatternFill(
                start_color=risk_colors.get(riesgo_f, "FFFFFF"),
                end_color=risk_colors.get(riesgo_f, "FFFFFF"),
                fill_type="solid"
            )
            
            for col_idx in range(1, 8):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = border_thin
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if col_idx == 2:  # Columna de riesgo
                    cell.fill = row_fill
                    cell.font = Font(bold=True)
            
            row += 1
    else:
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = "No se encontraron hallazgos para este paciente"
        cell.font = Font(italic=True, color="666666")
        cell.alignment = Alignment(horizontal='center')
        row += 1
    
    row += 2
    
    # ===== RECOMENDACIÓN GENERAL =====
    ws[f'A{row}'] = "RECOMENDACIÓN GENERAL"
    ws[f'A{row}'].font = title_font
    row += 1
    
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = audit_data.get("recomendacion_general", "Continuar con auditoría de rutina.")
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    cell.border = border_thin
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
